"""
Telegram bot for the AI PM Agent.
Handles text messages, file attachments, and inline keyboard button responses.
"""

import json
import logging
import os
import re
import tempfile
from pathlib import Path

from telegram import (
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    Update,
)
from telegram.constants import ChatAction, ParseMode
from telegram.ext import (
    Application,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    filters,
)

from agent import run_agent

TELEGRAM_TOKEN = os.environ["TELEGRAM_BOT_TOKEN"]
ALLOWED = {int(x) for x in os.environ.get("ALLOWED_CHAT_IDS", "").split(",") if x.strip()}
HISTORY_FILE = Path("/workspace/.bot_history.json")
MAX_HISTORY = 20
MAX_LEN = 4000

logging.basicConfig(format="%(asctime)s | %(levelname)s | %(message)s", level=logging.INFO)
logger = logging.getLogger("pm-bot")

history: dict[int, list] = {}

COMMAND_HINTS = {
    "project": [
        "Show all projects",
        "Create a new project for [name]",
        "Open project PROJ-001",
    ],
    "feature": [
        "Create a feature request for [description]",
        "Score FR-001 with RICE",
        "Gate review FR-001",
    ],
    "prd": [
        "Create PRD for FR-001",
        "Grill PRD-001",
        "Submit PRD-001-v1.0 for review",
    ],
    "epic": [
        "Create epic for PRD-001: [name]",
        "Approve EP-001-v1.0",
    ],
    "change": [
        "Create CR for PRD-001",
        "Assess CR-001",
        "Approve CR-001",
    ],
    "stakeholder": [
        "Add stakeholder [describe person]",
        "Add feedback from [name] - [feedback]",
        "Draft email to [name]",
    ],
}


# ── Inline keyboard helpers ───────────────────────────────────────────────────

def _extract_suggestions(text: str) -> list[str]:
    """
    Parse 'What to do next:' section from AI response.
    Returns list of command strings (up to 3).
    """
    pattern = r'(?:What to do next|Next steps?|Quick actions?):\s*\n((?:[-*]\s*"[^"]+".+\n?)+)'
    match = re.search(pattern, text, re.IGNORECASE)
    if not match:
        return []
    commands = []
    for line in match.group(1).split("\n"):
        q = re.search(r'"([^"]{3,60})"', line)
        if q:
            commands.append(q.group(1))
    return commands[:3]


def _strip_suggestion_block(text: str) -> str:
    """Remove the 'What to do next:' block from response text."""
    cleaned = re.sub(
        r"\n?---\s*\nWhat to do next:.*",
        "",
        text,
        flags=re.DOTALL | re.IGNORECASE,
    )
    return cleaned.strip()


def _make_keyboard(suggestions: list[str]) -> InlineKeyboardMarkup | None:
    if not suggestions:
        return None
    buttons = [
        [InlineKeyboardButton(f"-> {s}", callback_data=s[:64])]
        for s in suggestions
    ]
    return InlineKeyboardMarkup(buttons)


# ── Smart fallback ────────────────────────────────────────────────────────────

def _smart_fallback(user_message: str, hist: list, error_str: str = "") -> str:
    if any(w in error_str.lower() for w in ("quota", "rate limit", "429")):
        return (
            "The AI service is temporarily busy. Please wait 30 seconds and try again.\n\n"
            "---\nWhat to do next:\n"
            '- "Show all projects" - review your current work\n'
            '- "/help" - see all commands'
        )
    if any(w in error_str.lower() for w in ("decommissioned", "model not found")):
        return "The AI model needs to be updated by your admin. Try /reset."

    msg = user_message.lower()
    hints = []
    if any(w in msg for w in ("project", "proj")):
        hints = COMMAND_HINTS["project"]
    elif any(w in msg for w in ("feature", "fr-", "score", "rice", "gate")):
        hints = COMMAND_HINTS["feature"]
    elif any(w in msg for w in ("prd", "requirement")):
        hints = COMMAND_HINTS["prd"]
    elif any(w in msg for w in ("epic", "ep-", "user story")):
        hints = COMMAND_HINTS["epic"]
    elif any(w in msg for w in ("cr", "change")):
        hints = COMMAND_HINTS["change"]
    elif any(w in msg for w in ("stakeholder", "feedback", "email")):
        hints = COMMAND_HINTS["stakeholder"]

    if not hints and hist:
        last = " ".join(str(m.get("content", "")) for m in hist[-4:]).lower()
        if "proj-" in last:
            hints = COMMAND_HINTS["feature"][:2]
        elif "fr-" in last:
            hints = COMMAND_HINTS["prd"][:2]
        elif "prd-" in last:
            hints = COMMAND_HINTS["epic"][:2]

    if not hints:
        hints = [
            "Show all projects",
            "Create a new project for [name]",
            "Create a feature request for [description]",
        ]

    hints_text = "\n".join(f'- "{h}"' for h in hints)
    return (
        "I didn't quite understand that. Here are commands that might match:\n\n"
        f"---\nWhat to do next:\n{hints_text}"
    )


# ── File attachment conversion ────────────────────────────────────────────────

def _convert_file_to_md(file_path: Path, filename: str) -> str:
    suffix = file_path.suffix.lower()
    if suffix == ".md":
        return file_path.read_text(encoding="utf-8", errors="ignore")
    if suffix in (".docx", ".doc"):
        try:
            from docx import Document
            doc = Document(str(file_path))
            lines = []
            for para in doc.paragraphs:
                style = para.style.name.lower()
                text = para.text.strip()
                if not text:
                    lines.append("")
                elif "heading 1" in style:
                    lines.append(f"# {text}")
                elif "heading 2" in style:
                    lines.append(f"## {text}")
                elif "heading 3" in style:
                    lines.append(f"### {text}")
                else:
                    lines.append(text)
            return "\n".join(lines)
        except Exception as e:
            return f"[Could not read .docx: {e}]"
    if suffix == ".pdf":
        try:
            from pypdf import PdfReader
            reader = PdfReader(str(file_path))
            return "\n\n".join(p.extract_text() or "" for p in reader.pages)
        except Exception as e:
            return f"[Could not read .pdf: {e}]"
    if suffix in (".xlsx", ".xls"):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(file_path), read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True, max_row=20))
            if not rows:
                return "[Empty spreadsheet]"
            headers = [str(c) if c is not None else "" for c in rows[0]]
            md = "| " + " | ".join(headers) + " |\n"
            md += "| " + " | ".join(["---"] * len(headers)) + " |\n"
            for row in rows[1:]:
                md += "| " + " | ".join(str(c) if c is not None else "" for c in row) + " |\n"
            return md
        except Exception as e:
            return f"[Could not read spreadsheet: {e}]"
    if suffix == ".csv":
        try:
            text = file_path.read_text(encoding="utf-8", errors="ignore")
            lines = text.strip().split("\n")
            if not lines:
                return "[Empty CSV]"
            headers = lines[0].split(",")
            md = "| " + " | ".join(headers) + " |\n"
            md += "| " + " | ".join(["---"] * len(headers)) + " |\n"
            for row in lines[1:10]:
                md += "| " + " | ".join(row.split(",")) + " |\n"
            return md
        except Exception as e:
            return f"[Could not read CSV: {e}]"
    if suffix in (".txt", ".text"):
        return file_path.read_text(encoding="utf-8", errors="ignore")
    return f"[Unsupported file type: {suffix}]"


# ── Helpers ───────────────────────────────────────────────────────────────────

def _load():
    if HISTORY_FILE.exists():
        try:
            for k, v in json.loads(HISTORY_FILE.read_text()).items():
                history[int(k)] = v
        except Exception:
            pass


def _save():
    try:
        HISTORY_FILE.write_text(
            json.dumps({str(k): v for k, v in history.items()}, ensure_ascii=False)
        )
    except Exception:
        pass


def _allowed(chat_id):
    return not ALLOWED or chat_id in ALLOWED


def _split(text):
    if len(text) <= MAX_LEN:
        return [text]
    chunks = []
    while text:
        chunk = text[:MAX_LEN]
        nl = chunk.rfind("\n")
        if nl > MAX_LEN // 2:
            chunk = text[:nl]
        chunks.append(chunk)
        text = text[len(chunk):].lstrip("\n")
    return chunks


async def _send(update_or_query, ctx, text: str, chat_id: int):
    """Send text chunks + inline keyboard from suggestions."""
    suggestions = _extract_suggestions(text)
    clean = _strip_suggestion_block(text)
    keyboard = _make_keyboard(suggestions)

    chunks = _split(clean)
    for i, chunk in enumerate(chunks):
        kb = keyboard if (i == len(chunks) - 1) else None
        try:
            await ctx.bot.send_message(
                chat_id,
                chunk,
                parse_mode=ParseMode.MARKDOWN,
                reply_markup=kb,
            )
        except Exception:
            await ctx.bot.send_message(chat_id, chunk, reply_markup=kb)


# ── Handlers ──────────────────────────────────────────────────────────────────

async def cmd_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not _allowed(update.effective_chat.id):
        return
    text = (
        "PM Agent ready.\n\n"
        "Tap a button below or type your own command:\n\n"
        "---\nWhat to do next:\n"
        '- "Create a new project for [name]"\n'
        '- "Show all projects"\n'
        '- "Create a feature request for [description]"'
    )
    await _send(update, ctx, text, update.effective_chat.id)


async def cmd_reset(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not _allowed(update.effective_chat.id):
        return
    history.pop(update.effective_chat.id, None)
    _save()
    await update.message.reply_text("Conversation cleared. Ready for a fresh start.")


async def cmd_help(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not _allowed(update.effective_chat.id):
        return
    text = (
        "Available commands:\n\n"
        "Projects\n"
        '  "Show all projects"\n'
        '  "Create a new project for [name]"\n'
        '  "Open project PROJ-001"\n\n'
        "Features\n"
        '  "Create a feature request for [description]"\n'
        '  "Score FR-001 with RICE"\n'
        '  "Gate review FR-001"\n\n'
        "PRD\n"
        '  "Create PRD for FR-001"\n'
        '  "Grill PRD-001"\n'
        '  "Submit PRD-001-v1.0 for review"\n\n'
        "Epics\n"
        '  "Create epic for PRD-001: [name]"\n\n'
        "Change Requests\n"
        '  "Create CR for PRD-001"\n'
        '  "Assess CR-001" | "Approve CR-001"\n\n'
        "Stakeholders\n"
        '  "Add stakeholder [describe person]"\n'
        '  "Add feedback from [name] - [feedback]"\n'
        '  "Draft email to [name]"\n\n'
        "Files\n"
        "  Send any .docx .pdf .xlsx .md .csv file\n\n"
        "/reset - clear history | /start - welcome"
    )
    await update.message.reply_text(text)


async def handle(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handle text messages."""
    chat_id = update.effective_chat.id
    if not _allowed(chat_id) or not (update.message and update.message.text):
        return

    logger.info("chat_id=%d | %r", chat_id, update.message.text[:60])
    await ctx.bot.send_chat_action(chat_id=chat_id, action=ChatAction.TYPING)

    hist = history.setdefault(chat_id, [])
    user_text = update.message.text

    try:
        reply = await run_agent(user_text, hist)
    except Exception as e:
        logger.exception("Agent error: %s", e)
        reply = _smart_fallback(user_text, hist, str(e))

    if len(hist) > MAX_HISTORY:
        del hist[:len(hist) - MAX_HISTORY]
    _save()

    await _send(update, ctx, reply, chat_id)


async def handle_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handle inline keyboard button taps - treat as new message."""
    query = update.callback_query
    await query.answer()

    chat_id = query.message.chat_id
    if not _allowed(chat_id):
        return

    command = query.data
    logger.info("callback chat_id=%d | %r", chat_id, command)

    await ctx.bot.send_chat_action(chat_id=chat_id, action=ChatAction.TYPING)
    hist = history.setdefault(chat_id, [])

    try:
        reply = await run_agent(command, hist)
    except Exception as e:
        logger.exception("Agent error on callback: %s", e)
        reply = _smart_fallback(command, hist, str(e))

    if len(hist) > MAX_HISTORY:
        del hist[:len(hist) - MAX_HISTORY]
    _save()

    await _send(update, ctx, reply, chat_id)


async def handle_document(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handle file uploads."""
    chat_id = update.effective_chat.id
    if not _allowed(chat_id):
        return

    doc = update.message.document
    if not doc:
        return

    filename = doc.file_name or "uploaded_file"
    suffix = Path(filename).suffix.lower()
    supported = {".md", ".docx", ".doc", ".pdf", ".xlsx", ".xls", ".csv", ".txt"}

    if suffix not in supported:
        await update.message.reply_text(
            f"File type {suffix} is not supported.\n"
            "Supported: .md .docx .pdf .xlsx .csv .txt"
        )
        return

    await ctx.bot.send_chat_action(chat_id=chat_id, action=ChatAction.TYPING)
    await update.message.reply_text(f"Reading {filename}...")

    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp) / filename
        tg_file = await ctx.bot.get_file(doc.file_id)
        await tg_file.download_to_drive(str(tmp_path))
        content = _convert_file_to_md(tmp_path, filename)

    if len(content) > 8000:
        content = content[:8000] + "\n\n[... content truncated ...]"

    caption = update.message.caption or ""
    if caption:
        user_msg = f"File attached: {filename}\nInstruction: {caption}\n\nContent:\n{content}"
    else:
        user_msg = (
            f"File attached: {filename}\n\nContent:\n{content}\n\n"
            "---\nWhat to do next:\n"
            '- "Use this as research for a feature request"\n'
            '- "Import this as a PRD or requirements document"\n'
            '- "Summarize the key points from this file"'
        )

    hist = history.setdefault(chat_id, [])
    try:
        reply = await run_agent(user_msg, hist)
    except Exception as e:
        logger.exception("Agent error on file: %s", e)
        reply = _smart_fallback(filename, hist, str(e))

    if len(hist) > MAX_HISTORY:
        del hist[:len(hist) - MAX_HISTORY]
    _save()

    await _send(update, ctx, reply, chat_id)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    _load()
    app = Application.builder().token(TELEGRAM_TOKEN).build()

    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("reset", cmd_reset))
    app.add_handler(CommandHandler("help", cmd_help))
    app.add_handler(CallbackQueryHandler(handle_callback))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle))

    logger.info("Bot started. Allowed chats: %s", list(ALLOWED) if ALLOWED else "all")
    app.run_polling(drop_pending_updates=True)


if __name__ == "__main__":
    main()
