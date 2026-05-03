"""File attachment upload - stores in _attachments/YYYY-MM-DD/"""
import os
from datetime import date
from pathlib import Path
from fastapi import APIRouter, UploadFile, File, HTTPException
import subprocess

router = APIRouter()
WORKSPACE = Path(os.environ.get("WORKSPACE_PATH", "/workspace"))


def _convert_to_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".md":
        return path.read_text(encoding="utf-8", errors="ignore")
    if suffix in (".docx", ".doc"):
        try:
            from docx import Document
            doc = Document(str(path))
            lines = []
            for para in doc.paragraphs:
                style = para.style.name.lower()
                text = para.text.strip()
                if not text:
                    lines.append("")
                elif "heading 1" in style:
                    lines.append(f"# {text}")
                elif "heading 2" in style:
                    lines.append(f"## {text}")
                else:
                    lines.append(text)
            return "\n".join(lines)
        except Exception as e:
            return f"[Could not parse .docx: {e}]"
    if suffix == ".pdf":
        try:
            from pypdf import PdfReader
            reader = PdfReader(str(path))
            return "\n\n".join(p.extract_text() or "" for p in reader.pages)
        except Exception as e:
            return f"[Could not parse .pdf: {e}]"
    if suffix in (".xlsx", ".xls"):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(path), read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True, max_row=50))
            if not rows:
                return "[Empty spreadsheet]"
            headers = [str(c) if c is not None else "" for c in rows[0]]
            md = "| " + " | ".join(headers) + " |\n"
            md += "| " + " | ".join(["---"] * len(headers)) + " |\n"
            for row in rows[1:]:
                md += "| " + " | ".join(str(c) if c is not None else "" for c in row) + " |\n"
            return md
        except Exception as e:
            return f"[Could not parse spreadsheet: {e}]"
    if suffix == ".csv":
        try:
            text = path.read_text(encoding="utf-8", errors="ignore")
            lines = text.strip().split("\n")
            if not lines:
                return "[Empty CSV]"
            headers = lines[0].split(",")
            md = "| " + " | ".join(headers) + " |\n"
            md += "| " + " | ".join(["---"] * len(headers)) + " |\n"
            for row in lines[1:20]:
                md += "| " + " | ".join(row.split(",")) + " |\n"
            return md
        except Exception as e:
            return f"[Could not parse CSV: {e}]"
    return path.read_text(encoding="utf-8", errors="ignore")


@router.post("/upload")
async def upload_attachment(file: UploadFile = File(...)):
    today = date.today().isoformat()
    attach_dir = WORKSPACE / "_attachments" / today
    attach_dir.mkdir(parents=True, exist_ok=True)

    save_path = attach_dir / file.filename
    content_bytes = await file.read()
    save_path.write_bytes(content_bytes)

    # Convert to text for AI context
    text_content = _convert_to_text(save_path)
    if len(text_content) > 10000:
        text_content = text_content[:10000] + "\n\n[... truncated ...]"

    return {
        "filename": file.filename,
        "path": str(save_path.relative_to(WORKSPACE)),
        "content": text_content,
        "size": len(content_bytes),
    }


@router.get("/list")
async def list_attachments():
    attach_dir = WORKSPACE / "_attachments"
    if not attach_dir.exists():
        return {"attachments": []}
    files = []
    for f in sorted(attach_dir.rglob("*"), key=lambda x: x.stat().st_mtime, reverse=True):
        if f.is_file():
            files.append({
                "name": f.name,
                "path": str(f.relative_to(WORKSPACE)),
                "date": f.parent.name,
                "size": f.stat().st_size,
            })
    return {"attachments": files[:50]}
